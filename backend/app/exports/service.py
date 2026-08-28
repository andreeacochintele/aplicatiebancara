"""Business transaction export (architecture.md §25): a filtered CSV/XLSX of a
business user's own transaction activity, generated synchronously and
in-memory — same shape as statements/service.py's to_csv, no dedicated
export/job table for the *generation* step. See the QA punch-list report for
why this follows the statements precedent rather than architecture.md's
documented async exports-table model: nothing else in this codebase has
file-storage/background job infrastructure to build on, and "Format inițial:
CSV" (architecture.md §25) doesn't need it.

Each generated export is still logged to the `exports` table (ExportJob) so a
user can see and re-download past exports — see exports/models.py for how
that deviates from the doc's async-job shape."""
import base64
import csv
import io
import uuid
from datetime import datetime, time, timezone
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.business.repository import BusinessProfileRepository
from app.core.exceptions import NotFoundError, ValidationError
from app.exports.models import ExportFormat, ExportJob, ExportType
from app.exports.repository import ExportJobRepository, ExportRepository
from app.exports.schemas import (
    ExportCurrencyTotal,
    ExportedTransaction,
    ExportJobPublic,
    TransactionExportPreview,
    TransactionExportRequest,
)
from app.merchants.repository import MerchantRepository
from app.transactions.models import LedgerEntryType
from app.users.repository import UserRepository
from app.wallets.repository import WalletRepository

_CSV_COLUMNS = [
    "date", "transaction_id", "type", "counterparty", "description", "category",
    "direction", "amount", "currency", "status",
]

# Both the rendered file and the row-level preview are built fully in memory
# and (via generate_and_log) stored inline as a single DB column — unbounded
# like this, a business account with a long history could pull its entire
# transaction lifetime into one request/row. Capping the date range (like
# fx/service.py's rate-history endpoint caps `days`) is a first line of
# defense; row-count pagination would be the next step if this module ever
# grows past on-demand CSV/XLSX generation.
_MAX_EXPORT_RANGE_DAYS = 366


class ExportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.repository = ExportRepository(db)
        self.jobs = ExportJobRepository(db)
        self.wallets = WalletRepository(db)
        self.users = UserRepository(db)
        self.business_profiles = BusinessProfileRepository(db)
        self.merchants = MerchantRepository(db)

    def build_preview(self, user_id: uuid.UUID, data: TransactionExportRequest) -> TransactionExportPreview:
        if data.date_from > data.date_to:
            raise ValidationError("date_from must not be after date_to")
        if (data.date_to - data.date_from).days > _MAX_EXPORT_RANGE_DAYS:
            raise ValidationError(f"Date range too large — export at most {_MAX_EXPORT_RANGE_DAYS} days at a time")
        if data.wallet_id is not None:
            wallet = self.wallets.get_by_id(data.wallet_id)
            if wallet is None or wallet.user_id != user_id:
                raise NotFoundError("Wallet not found")

        period_start = datetime.combine(data.date_from, time.min, tzinfo=timezone.utc)
        period_end = datetime.combine(data.date_to, time.max, tzinfo=timezone.utc)

        entries = self.repository.list_entries_for_user(
            user_id,
            period_start,
            period_end,
            wallet_id=data.wallet_id,
            currency=data.currency,
            direction=data.direction,
            status=data.status,
            category_id=data.category_id,
        )

        category_ids = {e.transaction.category_id for e in entries if e.transaction.category_id is not None}
        category_names = self.repository.get_category_names(category_ids)

        counterparty_cache: dict[tuple[str, uuid.UUID], str] = {}
        transactions = [
            ExportedTransaction(
                date=entry.created_at,
                transaction_id=entry.transaction_id,
                type=entry.transaction.type,
                counterparty=self._resolve_counterparty(entry.transaction, counterparty_cache),
                description=entry.transaction.description,
                category=category_names.get(entry.transaction.category_id),
                direction="IN" if entry.entry_type == LedgerEntryType.CREDIT else "OUT",
                amount=entry.amount,
                currency=entry.currency,
                status=entry.transaction.status,
            )
            for entry in entries
        ]

        totals: dict[str, dict[str, Decimal]] = {}
        for entry in entries:
            bucket = totals.setdefault(entry.currency, {"in": Decimal("0"), "out": Decimal("0")})
            if entry.entry_type == LedgerEntryType.CREDIT:
                bucket["in"] += entry.amount
            elif entry.entry_type == LedgerEntryType.DEBIT:
                bucket["out"] += entry.amount

        return TransactionExportPreview(
            date_from=data.date_from,
            date_to=data.date_to,
            row_count=len(transactions),
            totals=[
                ExportCurrencyTotal(currency=currency, total_incoming=v["in"], total_outgoing=v["out"])
                for currency, v in sorted(totals.items())
            ],
            transactions=transactions,
        )

    def generate_and_log(
        self, user_id: uuid.UUID, data: TransactionExportRequest, file_format: ExportFormat
    ) -> tuple[ExportJob, bytes, str]:
        """Build the preview, render it to the requested format, and log an
        ExportJob row so it shows up in export history — used by both the
        direct-download route and (later) any route that just wants the
        history entry."""
        row_count: int
        if file_format == ExportFormat.MT940:
            # MT940 ignores every filter but wallet_id/dates (see to_mt940's
            # docstring), so its row count comes from the statement it
            # actually rendered, not the filtered `preview` below.
            text = self.to_mt940(user_id, data)
            content = text.encode("utf-8")
            media_type = "application/octet-stream"
            extension = "sta"
            stored_content = text
            row_count = text.count(":61:")
        else:
            preview = self.build_preview(user_id, data)
            row_count = preview.row_count
            if file_format == ExportFormat.XLSX:
                content = self.to_xlsx(preview)
                media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                extension = "xlsx"
                stored_content = base64.b64encode(content).decode("ascii")
            elif file_format == ExportFormat.PDF:
                active_profile = next((p for p in self.business_profiles.list_for_user(user_id) if p.is_active), None)
                content = self.to_pdf(
                    preview,
                    company_name=active_profile.company_name if active_profile else None,
                    representative_name=active_profile.representative_name if active_profile else None,
                )
                media_type = "application/pdf"
                extension = "pdf"
                stored_content = base64.b64encode(content).decode("ascii")
            else:
                content = self.to_csv(preview).encode("utf-8")
                media_type = "text/csv"
                extension = "csv"
                stored_content = content.decode("utf-8")

        job = self.jobs.add(
            ExportJob(
                user_id=user_id,
                type=ExportType.BUSINESS_TRANSACTIONS,
                format=file_format,
                date_from=data.date_from,
                date_to=data.date_to,
                filters={
                    k: (str(v) if not isinstance(v, str) else v)
                    for k, v in data.model_dump(exclude={"date_from", "date_to"}).items()
                    if v is not None
                },
                row_count=row_count,
                content=stored_content,
            )
        )
        self.db.commit()
        filename = f"transactions_{data.date_from}_{data.date_to}.{extension}"
        return job, content, f"{media_type}|{filename}"

    def list_history(self, user_id: uuid.UUID) -> list[ExportJobPublic]:
        return [ExportJobPublic.model_validate(job, from_attributes=True) for job in self.jobs.list_for_user(user_id)]

    _BINARY_FORMATS = {
        ExportFormat.XLSX: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ExportFormat.PDF: "application/pdf",
    }

    def download_job(self, user_id: uuid.UUID, job_id: uuid.UUID) -> tuple[ExportJob, bytes, str]:
        job = self.jobs.get_owned_by_id(user_id, job_id)
        if job is None or job.content is None:
            raise NotFoundError("Export not found")
        extension = "sta" if job.format == ExportFormat.MT940 else job.format.value.lower()
        media_type = self._BINARY_FORMATS.get(job.format, "text/csv" if job.format == ExportFormat.CSV else "application/octet-stream")
        content = base64.b64decode(job.content) if job.format in self._BINARY_FORMATS else job.content.encode("utf-8")
        filename = f"transactions_{job.date_from}_{job.date_to}.{extension}"
        return job, content, f"{media_type}|{filename}"

    def _resolve_counterparty(self, transaction, cache: dict[tuple[str, uuid.UUID], str]) -> str:
        if transaction.merchant_id is not None:
            key = ("merchant", transaction.merchant_id)
            if key not in cache:
                merchant = self.merchants.get_by_id(transaction.merchant_id)
                cache[key] = merchant.name if merchant is not None else ""
            return cache[key]
        if transaction.counterparty_user_id is not None:
            key = ("user", transaction.counterparty_user_id)
            if key not in cache:
                counterparty = self.users.get_by_id(transaction.counterparty_user_id)
                cache[key] = f"{counterparty.first_name} {counterparty.last_name}" if counterparty is not None else ""
            return cache[key]
        return ""

    @staticmethod
    def to_csv(preview: TransactionExportPreview) -> str:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(_CSV_COLUMNS)
        for tx in preview.transactions:
            writer.writerow(
                [
                    tx.date.isoformat(),
                    str(tx.transaction_id),
                    tx.type.value,
                    tx.counterparty,
                    tx.description or "",
                    tx.category or "",
                    tx.direction,
                    tx.amount,
                    tx.currency,
                    tx.status.value,
                ]
            )
        if preview.totals:
            writer.writerow([])
            writer.writerow(["TOTALS", "currency", "incoming", "outgoing"])
            for total in preview.totals:
                writer.writerow(["", total.currency, total.total_incoming, total.total_outgoing])
        return buffer.getvalue()

    @staticmethod
    def to_xlsx(preview: TransactionExportPreview) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Transactions"

        # Light branding: a title block above the data table. No embedded
        # logo image here (that needs Pillow via openpyxl, not currently a
        # dependency) — just styled title cells, kept out of the data rows
        # below so nothing that imports this sheet has to skip a header.
        sheet["A1"] = "EasyB"
        sheet["A1"].font = Font(bold=True, size=16, color="5B5FEF")
        sheet["A2"] = "Business Transaction Export"
        sheet["A2"].font = Font(size=10, color="6C6C82")
        sheet["A3"] = f"{preview.date_from} to {preview.date_to}"
        sheet["A3"].font = Font(size=9, color="6C6C82")

        header_row = 5
        for col_index, column in enumerate(_CSV_COLUMNS, start=1):
            cell = sheet.cell(row=header_row, column=col_index, value=column)
            cell.font = Font(bold=True)
        for tx in preview.transactions:
            sheet.append(
                [
                    tx.date.replace(tzinfo=None),
                    str(tx.transaction_id),
                    tx.type.value,
                    tx.counterparty,
                    tx.description or "",
                    tx.category or "",
                    tx.direction,
                    float(tx.amount),
                    tx.currency,
                    tx.status.value,
                ]
            )
        for index, column in enumerate(_CSV_COLUMNS, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = max(12, len(column) + 2)

        if preview.totals:
            totals_sheet = workbook.create_sheet("Totals")
            totals_sheet.append(["currency", "incoming", "outgoing"])
            for cell in totals_sheet[1]:
                cell.font = Font(bold=True)
            for total in preview.totals:
                totals_sheet.append([total.currency, float(total.total_incoming), float(total.total_outgoing)])

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def to_pdf(
        preview: TransactionExportPreview, company_name: str | None = None, representative_name: str | None = None
    ) -> bytes:
        from app.core.pdf_branding import BORDER, GREEN, RED, ROW_ALT, TEXT_DARK, TEXT_SOFT, gradient_color, new_branded_pdf

        pdf = new_branded_pdf(subtitle="Business Transaction Export", orientation="L")
        pdf.footer_note = "sandbox export, not a legal document"

        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(*TEXT_DARK)
        pdf.cell(0, 8, company_name or "Transaction export", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9.5)
        pdf.set_text_color(*TEXT_SOFT)
        if representative_name:
            pdf.cell(0, 6, f"Representative: {representative_name}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 6, f"Period: {preview.date_from} to {preview.date_to}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        if preview.totals:
            col_w = (28, 45, 45)
            pdf.set_font("Helvetica", "B", 8.5)
            pdf.set_fill_color(*gradient_color(0.15))
            pdf.set_text_color(255, 255, 255)
            for header, width in zip(("Currency", "Total incoming", "Total outgoing"), col_w):
                pdf.cell(width, 7, header, fill=True)
            pdf.ln()
            pdf.set_font("Helvetica", "", 8.5)
            for i, total in enumerate(preview.totals):
                pdf.set_fill_color(*(ROW_ALT if i % 2 else (255, 255, 255)))
                pdf.set_text_color(*TEXT_DARK)
                pdf.cell(col_w[0], 7, total.currency, border="B", fill=True)
                pdf.set_text_color(*GREEN)
                pdf.cell(col_w[1], 7, f"+{total.total_incoming}", border="B", fill=True)
                pdf.set_text_color(*RED)
                pdf.cell(col_w[2], 7, f"-{total.total_outgoing}", border="B", fill=True)
                pdf.ln()
            pdf.ln(5)

        col_widths = (22, 26, 35, 50, 24, 16, 22, 16, 22)
        headers = ("Date", "Type", "Counterparty", "Description", "Category", "Direction", "Amount", "Currency", "Status")
        pdf.set_font("Helvetica", "B", 8.5)
        pdf.set_fill_color(*gradient_color(0.15))
        pdf.set_text_color(255, 255, 255)
        for header, width in zip(headers, col_widths):
            pdf.cell(width, 8, header, fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        if not preview.transactions:
            pdf.set_text_color(*TEXT_SOFT)
            pdf.cell(sum(col_widths), 10, "No transactions match these filters.", border="LRB")
            pdf.ln()
        for i, tx in enumerate(preview.transactions):
            pdf.set_fill_color(*(ROW_ALT if i % 2 else (255, 255, 255)))
            pdf.set_draw_color(*BORDER)
            row = (
                tx.date.strftime("%Y-%m-%d"),
                tx.type.value.replace("_", " ").title(),
                tx.counterparty[:24],
                (tx.description or "")[:36],
                (tx.category or "")[:18],
                tx.direction,
                f"{'+' if tx.direction == 'IN' else '-'}{tx.amount}",
                tx.currency,
                tx.status.value,
            )
            pdf.set_text_color(*TEXT_DARK)
            for col, (value, width) in enumerate(zip(row, col_widths)):
                if col == 6:
                    pdf.set_text_color(*(GREEN if tx.direction == "IN" else RED))
                pdf.cell(width, 7, value, border="B", fill=True)
                if col == 6:
                    pdf.set_text_color(*TEXT_DARK)
            pdf.ln()

        return bytes(pdf.output())

    def to_mt940(self, user_id: uuid.UUID, data: TransactionExportRequest) -> str:
        """MT940 (SWIFT customer statement) is inherently a per-account,
        unfiltered statement — a running balance only reconciles against the
        account's complete ledger, so this ignores every business-export
        filter except wallet_id/date_from/date_to and reuses
        StatementService's own opening/closing-balance computation rather
        than re-deriving it here."""
        if data.wallet_id is None:
            raise ValidationError("MT940 export requires a specific wallet_id (MT940 is a per-account statement)")

        from app.statements.schemas import StatementRequest
        from app.statements.service import StatementService

        statement = StatementService(self.db).generate(
            user_id, StatementRequest(wallet_id=data.wallet_id, date_from=data.date_from, date_to=data.date_to)
        )

        def amount(value: Decimal) -> str:
            # SWIFT amounts: comma decimal separator, no thousands separator,
            # no sign (the D/C marker in front carries it).
            return f"{abs(value):.2f}".replace(".", ",")

        lines: list[str] = []
        lines.append(f":20:{statement.wallet_id.hex[:16].upper()}")
        lines.append(f":25:{statement.iban}")
        lines.append(":28C:1/1")  # no persistent statement sequence numbering exists yet — always page 1 of 1
        opening_dc = "D" if statement.opening_balance < 0 else "C"
        lines.append(
            f":60F:{opening_dc}{statement.date_from.strftime('%y%m%d')}{statement.currency}{amount(statement.opening_balance)}"
        )
        for tx in statement.transactions:
            value_date = tx.created_at.strftime("%y%m%d")
            entry_date = tx.created_at.strftime("%m%d")
            dc = "C" if tx.direction == "IN" else "D"
            lines.append(f":61:{value_date}{entry_date}{dc}{amount(tx.amount)}NTRFNONREF")
            narrative = (tx.description or tx.type.value.replace("_", " ").title())[:65]
            lines.append(f":86:{narrative}")
        closing_dc = "D" if statement.closing_balance < 0 else "C"
        lines.append(
            f":62F:{closing_dc}{statement.date_to.strftime('%y%m%d')}{statement.currency}{amount(statement.closing_balance)}"
        )
        lines.append("-")
        return "\r\n".join(lines) + "\r\n"
